from django.contrib.auth.models import User
from .models import Task, Role, Time, Project, Client, ProjectTask, UserProfile
from django.http import HttpResponse
import time

def dictfetchall(cursor):
    # Return all rows from a cursor as a dict
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

def get_table(project_id, start_date, end_date):
  from django.db import connection

  cursor = connection.cursor()

  sql = """select p.name as project_name, CONCAT(u.first_name, ' ', u.last_name) as whole_name, task.name as task_name, SUM(time.duration) as duration_sum
          from itimetracker_time as time
          inner join itimetracker_projecttask as pt
          on time.project_task_id = pt.id
          inner join itimetracker_project as p
          on p.id = pt.project_id
          inner join itimetracker_task as task
          on pt.task_id = task.id
          inner join itimetracker_userprofile as up
          on up.user_id = pt.user_id
          inner join auth_user as u
          on u.id = up.user_id
          inner join itimetracker_role as r
          on up.role_id = r.id
          where pt.project_id = %d""" % project_id
          
  if start_date != "" and end_date != "":
    sql += " AND time.date BETWEEN '%s' AND '%s'" % (start_date, end_date)

  sql += " GROUP BY p.name, u.first_name, u.last_name, task.name, u.id, p.id, task.id ORDER BY u.id, p.id, task.id"

  cursor.execute(sql)

  dic = dictfetchall(cursor)  
  user_count = 0 
  user_dic = {}


  # create dictionary ordered by employee
  for d in dic:
    wn = d["whole_name"]
    tn = d["task_name"]
    ds = d["duration_sum"]
    if wn not in user_dic:
      user_count += 1
      user_dic[wn] = []
    user_dic[wn].append([tn, ds])

  table = [[" "]]
  tasks = [] # list of tasks on project. See users list.

  task_totals = [] # total per row
  user_totals = [] # total per column

  i = 0
  for name,task_list in user_dic.iteritems():
    i += 1 # user index
    table[0].append(name) # append employee full name to column header
    # j is a task counter per employee.. maybe not needed
    user_total = 0
    for j,tl in enumerate(task_list): # for the given employee's list of tasks
      if tl[0] not in tasks: # if the task isn't included in the table yet..
        tasks.append(tl[0]) # add it to the task list for memorizing purposes
        task_totals.append(tl[1])
        table.append([tl[0]]) # add a row to the table
        for x in range(1, user_count + 1): # fill in the row with the value and 0's
          if x is not i: # not current user's column
            table[-1].append(0)
          else: # is current user's column
            table[-1].append(tl[1])
            user_total += tl[1]
      else: # task is already included in the table
        # match the index in the row to the correct user's column
        row = tasks.index(tl[0]) + 1
        col = i
        table[row][col] += tl[1]
        user_total += tl[1]
        task_totals[row-1] += tl[1]
    user_totals.append(user_total)
    

  table[0].append('Totals')
  table.append(['Totals'])

  for i,task_total in enumerate(task_totals):
    table[i+1].append(task_total)

  for user_total in user_totals:
    table[-1].append(user_total)
  table[-1].append(' ')

  return table

def export_xlsx(project_id_list, start_date, end_date):
  from openpyxl import Workbook
  from openpyxl.compat import range
  from openpyxl.cell import get_column_letter

  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  #response['Content-Disposition'] = 'attachment; filename=' + start_date.strftime('%Y%m%d') + '_to_' + end_date.strftime('%Y%m%d') + '.xlsx'
  response['Content-Disposition'] = 'attachment; filename=project_exports.xlsx'

  wb = Workbook()
  ws = wb.active

  for i,project_id in enumerate(project_id_list):
    project_id = int(project_id)
    project = Project.objects.get(id=project_id)
    table = get_table(project_id, start_date, end_date)
    
    if i is not 0:
      ws = wb.create_sheet()
    ws.title = project.name[:31]

    # CODE IS REVIEWED AND ADJUSTED UP TO THIS POINT

    for row in range(1,len(table)+1):
      for col in range(1,len(table[0])+1):
        if col == 1 or row == 1 or (row == len(table) and col == len(table[row-1])):
          ws.cell(column=col, row=row, value="%s" % table[row-1][col-1])
        else:
          ws.cell(column=col, row=row).value = float(table[row-1][col-1])

  wb.save(response)

  return response

export_xlsx.short_description = u'Export Project to XLSX'
